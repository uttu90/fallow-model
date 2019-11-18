import os

import openpyxl as ox
import numpy as np

# file_path = os.path.dirname(__file__)
#
# data_file = os.path.join(file_path, 'fallow_data 210919 (7).xlsx')

ALPHABET = 'ABCDEFGHIJKLMNOPQRSTUVXYZ'


def read_data_as_array(ws, col, min_row, max_row):
    col_index = ALPHABET.index(col) + 1
    return np.array(
        list(filter(
            lambda x: x is not None,
            list(ws.iter_cols(
                min_col=col_index,
                max_col=col_index,
                min_row=min_row,
                max_row=max_row,
                values_only=True
            ))[0]))
    )


def read_ts_as_array(ws, min_col, max_col, min_row, max_row):
    return np.array(list(ws.iter_cols(
        min_col=min_col,
        max_col=max_col,
        min_row=min_row,
        max_row=max_row
    )))


def stat(mean, cv):
    return mean * (1 + np.random.normal(size=mean.shape) * cv)


def load_data(path):
    wb = ox.load_workbook(path, data_only=True)
    ws = wb['Summary']
    # es = wb['Economic']

    data = dict()

    data['rotation'] = read_data_as_array(ws, 'D', 6, 20)
    data['allow_change'] = read_data_as_array(ws, 'E', 6, 20)
    data['landcover_age_boundary'] = read_data_as_array(ws, 'B', 27, 76)
    data['initial_landcover_age_mean'] = read_data_as_array(ws, 'C', 27, 76)
    data['initial_landcover_age_cv'] = read_data_as_array(ws, 'D', 27, 76)
    data['depletion_rate_mean'] = read_data_as_array(ws, 'E', 27, 76)
    data['depletion_rate_cv'] = read_data_as_array(ws, 'F', 27, 76)
    data['halftime_recovery_mean'] = read_data_as_array(ws, 'G', 27, 76)
    data['halftime_recovery_cv'] = read_data_as_array(ws, 'H', 27, 76)
    data['aboveground_biomass_mean'] = read_data_as_array(ws, 'I', 27, 76)
    data['aboveground_biomass_cv'] = read_data_as_array(ws, 'J', 27, 76)
    data['floor_biomass_fraction_mean'] = read_data_as_array(ws, 'K', 27, 76)
    data['floor_biomass_fraction_cv'] = read_data_as_array(ws, 'L', 27, 76)
    data['land_yield_mean'] = read_data_as_array(ws, 'M', 27, 76)
    data['land_yield_cv'] = read_data_as_array(ws, 'N', 27, 76)
    data['probably_of_fire_mean'] = read_data_as_array(ws, 'O', 27, 76)
    data['probably_of_fire_cv'] = read_data_as_array(ws, 'P', 27, 76)
    data['harvesting_production_mean'] = read_data_as_array(ws, 'B', 82, 96)
    data['harvesting_production_cv'] = read_data_as_array(ws, 'C', 82, 96)
    data['demand_per_capital'] = read_data_as_array(ws, 'D', 82, 96)
    data['probably_to_sell'] = read_data_as_array(ws, 'E', 82, 96)
    data['soil_fertility'] = read_data_as_array(ws, 'F', 82, 96)
    data['land_productivity'] = read_data_as_array(ws, 'G', 82, 96)
    data['land_suitability'] = read_data_as_array(ws, 'H', 82, 96)
    data['transportation_access'] = read_data_as_array(ws, 'I', 82, 96)
    data['plot_maintainance'] = read_data_as_array(ws, 'J', 82, 96)
    data['slope'] = read_data_as_array(ws, 'K', 82, 96)
    data['floor_biomass'] = read_data_as_array(ws, 'L', 82, 96)
    data['loss_fraction'] = read_data_as_array(ws, 'N', 82, 96)
    data['price_mean'] = read_data_as_array(ws, 'B', 103, 117)
    data['price_cv'] = read_data_as_array(ws, 'C', 103, 117)
    data['return_to_land_farmer_1'] = read_data_as_array(ws, 'D', 103, 117)
    data['return_to_land_farmer_2'] = read_data_as_array(ws, 'E', 103, 117)
    data['init_return_to_land'] = np.dstack((data['return_to_land_farmer_1'], data['return_to_land_farmer_2']))[0]
    data['return_to_labour_farmer_1'] = read_data_as_array(ws, 'F', 103, 117)
    data['return_to_labour_farmer_2'] = read_data_as_array(ws, 'G', 103, 117)
    data['init_return_to_labour'] = np.dstack((data['return_to_labour_farmer_1'], data['return_to_labour_farmer_2']))[0]
    data['expect_return_to_land'] = read_data_as_array(ws, 'H', 103, 117)
    data['expect_return_to_labour'] = read_data_as_array(ws, 'I', 103, 117)
    data['establishment_cost_mean'] = read_data_as_array(ws, 'J', 103, 117)
    data['establishment_cost_cv'] = read_data_as_array(ws, 'K', 103, 117)
    data['establishment_labour_mean'] = read_data_as_array(ws, 'L', 103, 117)
    data['establishment_labour_cv'] = read_data_as_array(ws, 'M', 103, 117)
    data['external_labour_mean'] = read_data_as_array(ws, 'N', 103, 117)
    data['external_labour_cv'] = read_data_as_array(ws, 'O', 103, 117)
    data['subsidy_mean'] = read_data_as_array(ws, 'P', 103, 117)
    data['subsidy_cv'] = read_data_as_array(ws, 'Q', 103, 117)
    data['non_labour_cost_mean'] = read_data_as_array(ws, 'B', 121, 171)
    data['non_labour_cost_cv'] = read_data_as_array(ws, 'C', 121, 171)
    data['cultural_influence'] = read_data_as_array(ws, 'B', 177, 191)
    data['availability'] = read_data_as_array(ws, 'C', 177, 191)
    data['credibility'] = read_data_as_array(ws, 'D', 177, 191)
    data['demography'] = read_data_as_array(ws, 'C', 196, 201)
    data['farmer_prop_1'] = read_data_as_array(ws, 'C', 206, 209)
    data['farmer_prop_2'] = read_data_as_array(ws, 'D', 206, 209)
    data['agentprop'] = np.dstack((data['farmer_prop_1'], data['farmer_prop_2']))[0]
    data['impact_of_disaster'] = read_data_as_array(ws, 'B', 213, 215)
    data['time_of_disaster'] = read_data_as_array(ws, 'B', 216, 216)[0]
    data['conversion'] = read_data_as_array(ws, 'B', 218, 219)
    data['subsidy_est'] = read_data_as_array(ws, 'B', 284, 298)
    data['subsidy_mnt'] = read_data_as_array(ws, 'C', 284, 298)
    data['ex'] = read_ts_as_array(ws, 2, 32, 245, 259)
    data['sub'] = read_ts_as_array(ws, 2, 32, 264, 278)
    data['price_'] = read_ts_as_array(ws, 2, 32, 226, 240)
    data['explabor'] = read_data_as_array(ws, 'B', 303, 317)
    data['expland'] = read_data_as_array(ws, 'C', 303, 317)

    return data
