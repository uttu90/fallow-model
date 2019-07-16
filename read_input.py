from openpyxl import load_workbook

from constant import *

FILE_PATH = 'fallow_data.xlsx'


def create_data_object(attrs, data={}):
  for attr in attrs:
    if (type(attr) is str):
      data[attr] = {}
    else:
      key = list(attr)[0]
      data[key] = {}
      create_data_object(attr[key], data[key])


def assign_headers_value(attrs, values):
  data = {}
  def assign_value(attrs, values, data={}):
    for attr in attrs:
      if (type(attr) is str):
        data[attr] = values[assign_value.index]
        assign_value.index = assign_value.index + 1
      else:
        for key in attr.keys():
          data[key] = {}
          assign_value(attr[key], values, data[key])
        # key = attr.keys()[0]
        # data[key] = {}
        # assign_value(attr[key], values, data[key])
  assign_value.index = 0
  assign_value(attrs, values, data)
  return data


def assign_columns_value(attrs, values):
  data = {}
  def assign_value(attrs, values, data={}):
    for attr in attrs:
      if (type(attr) is str):
        data[attr] = values[assign_value.index]
        assign_value.index = assign_value.index + 1
      else:
        key = list(attr)[0]
        data[key] = {}
        assign_value.index = assign_value.index + 1
        assign_value(attr[key], values, data[key])
  assign_value.index = 0
  assign_value(attrs, values, data)
  return data


def get_table_data(ws, min_row, max_row, start_col, end_col):
  min_col = ALPHABET.index(start_col) + 1
  max_col = ALPHABET.index(end_col) + 1
  if (min_col == max_col):
    return [
      row[0].value or 0
      for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
    ]
  return [
    [cell.value or 0 for cell in row]
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
  ]


def get_data_from_table(headers, columns, table):
  row_data = [
      assign_headers_value(headers, row)
      for row in table
  ]
  return assign_columns_value(columns, row_data)

BIOPHYSICAL1_TABLE = [27, 76, 'B', 'P']
BIOPHISICAL2_TABLE = [82, 96, 'B', 'Q']
ECONOMIC1_TABLE = [103, 117, 'B', 'Q']
ECONOMIC2_TABLE = [121, 171, 'B', 'C']
SOCIAL_CULTURE_TABLE = [177, 191, 'B', 'D']
IMPACT_TABLE = [212, 219, 'B', 'B']
DEMOGRAPHY_TABLE = [196, 201, 'C', 'C']
FARMER_TABLE = [206, 209, 'C', 'D']
SUBSIDY_TABLE = [285, 299, 'B', 'C']

PRODUCT_PRICE_TABLE = [226, 240, 'B', 'CW']
EXTENSION_AVAILABILITY_TABLE = [245, 259, 'B', 'CW']
SUBSIDY_AVAILABILITY_TABLE = [264, 278, 'B', 'CW']


def get_data(file=None):
    data = dict()
    ts = dict()

    wb = load_workbook(filename=file or FILE_PATH, data_only=True)

    summary = wb.get_sheet_by_name('Summary')

    data['land_specs'] = get_data_from_table(LAND_SPECS, LIVELIHOOD,
                                     get_table_data(summary,
                                                    *LAND_SPECS_TABLE))
    data['biophysical1'] = get_data_from_table(BIOPHYSICAL1, LANDCOVER,
                                       get_table_data(summary,
                                                      *BIOPHYSICAL1_TABLE))
    print(data['biophysical1'])
    data['biophysical2'] = get_data_from_table(BIOPHISICAL2, LIVELIHOOD,
                                       get_table_data(summary,
                                                      *BIOPHISICAL2_TABLE))
    data['economic1'] = get_data_from_table(
        ECONOMIC, LIVELIHOOD, get_table_data(summary, *ECONOMIC1_TABLE))
    data['economic2'] = get_data_from_table(
        ECONOMIC2, LANDCOVER, get_table_data(summary, *ECONOMIC2_TABLE))
    data['social'] = get_data_from_table(
        SOCIAL, LIVELIHOOD, get_table_data(summary, *SOCIAL_CULTURE_TABLE))
    data['farmer'] = get_data_from_table(FARMERS, FARMER_PROPERTY,
                                 get_table_data(summary, *FARMER_TABLE))
    data['demography'] = assign_columns_value(DEMOGRAPHY, get_table_data(summary,
                                                                 *DEMOGRAPHY_TABLE))
    data['impact'] = assign_columns_value(IMPACT,
                                  get_table_data(summary, *IMPACT_TABLE))
    data['subsidy'] = get_data_from_table(SUBSIDY_PARAMETER, LIVELIHOOD,
                                  get_table_data(summary, *SUBSIDY_TABLE))

    ts['product_price'] = assign_columns_value(LIVELIHOOD, get_table_data(summary,
                                                                    *PRODUCT_PRICE_TABLE))
    ts['extension_availability'] = assign_columns_value(LIVELIHOOD,
                                                  get_table_data(summary,
                                                                 *EXTENSION_AVAILABILITY_TABLE))
    ts['subsidy_availability'] = assign_columns_value(LIVELIHOOD,
                                                get_table_data(summary,
                                                               *SUBSIDY_AVAILABILITY_TABLE))
    return (data, ts)

# get_data(FILE_PATH)
# print(get_data(FILE_PATH))
